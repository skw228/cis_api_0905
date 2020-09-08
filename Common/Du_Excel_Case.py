from openpyxl import load_workbook
from Common.Read_Config import Readconfig
from Common.Project_Path import *


class Du_Excel_Case():
    def get_case(self, file_name):  # 读EXcel测试用例
        wb = load_workbook(file_name)
        title_name = eval(Readconfig.get_config(case_config_path, "TITLE", "title"))
        title = []
        for sheet_name in title_name:
            sheet = wb[sheet_name]
            for j in range(1, sheet.max_column + 1):
                title.append(sheet.cell(1, j).value)
        test_case = []
        mode = eval(Readconfig.get_config(case_config_path, "MODE", "mode"))
        for sheet_name in mode:
            sheet = wb[sheet_name]
            if mode[sheet_name] == "all":
                for i in range(3, sheet.max_row + 1):
                    data = {}
                    for j in range(1, sheet.max_column + 1):
                        if title[j - 1] != None:
                            data[title[j - 1]] = sheet.cell(i, j).value
                    test_case.append(data)
            else:
                for i in range(3, sheet.max_row + 1):
                    data = {}
                    for j in range(1, sheet.max_column + 1):
                        if title[j - 1] != None:
                            data[title[j - 1]] = sheet.cell(i + 1, j).value
                    test_case.append(data)
        return test_case

    def get_variable(self, file_name):  # 读EXcel变量基础数据
        wb = load_workbook(file_name)
        mode = eval(Readconfig.get_config(case_config_path, "MODE1", "mode1"))
        titie = []
        for sheet_name in mode:
            sheet = wb[sheet_name]
            for j in range(1, sheet.max_row + 1):
                titie.append(sheet.cell(j, 1).value)
        for sheet_name in mode:
            sheet = wb[sheet_name]
            data = {}
            for j in range(1, sheet.max_row + 1):
                if titie[j - 1] != None:
                    data[titie[j - 1]] = sheet.cell(j, 2).value
            return data


if __name__ == '__main__':
    res = Du_Excel_Case().get_case(test_case_path)
    print(res)
