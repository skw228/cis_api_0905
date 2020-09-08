from openpyxl import load_workbook
from Common.Read_Config import Readconfig
from Common.Get_Data import get_data


class Write_Excel_Data():
    def set_data(self, config_name, file_name):  # 变量结果写入
        variable_data = getattr(get_data, "variable_data")
        wb = load_workbook(file_name)
        mode = eval(Readconfig.get_config(config_name, "MODE1", "mode1"))
        for sheet_name in mode:
            sheet_var = wb[sheet_name]
            i = 0
            for key in variable_data.keys():
                i = i + 1
                sheet_var.cell(i, 1).value = key
                sheet_var.cell(i, 2).value = variable_data[key]
        mode = eval(Readconfig.get_config(config_name, "MODE", "mode"))
        for sheet_name in mode:
            sheet_case = wb[sheet_name]
            # 响应结果、执行结果写回
            request = get_data.assertion_data
            for key_request in request:
                for key in key_request.keys():
                    sheet_case.cell(int(key), 8).value = (key_request[key])[0]
                    sheet_case.cell(int(key), 9).value = (key_request[key])[1]
            # SQL查询结果写回
            sql_Request = get_data.sql_assertion_Request
            for key_sql_Request in sql_Request:
                for key in key_sql_Request.keys():
                    sheet_case.cell(int(key), 13).value = key_sql_Request[key]
            # SQL执行结果写回
            sql_Result = get_data.sql_assertion_Result
            for key_sql_Result in sql_Result:
                for key in key_sql_Result.keys():
                    sheet_case.cell(int(key), 14).value = key_sql_Result[key]
        wb.save(file_name)
